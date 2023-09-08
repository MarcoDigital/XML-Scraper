import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
import time

URL = "https://www.XXX.nl/sitemap.xml"
EXCEL_FILE_NAME = "extracted_urls.xlsx"
SLEEP_INTERVAL = 4 * 3600  # 4 uur slaap

def current_timestamp():
    return time.strftime('%Y-%m-%d %H:%M:%S')

def fetch_xml_content(url):
    response = requests.get(url)
    return response.content

def extract_urls_and_lastmod(xml_content):
    soup = BeautifulSoup(xml_content, 'lxml-xml')
    urls_data = []
    for url_tag in soup.find_all('url'):
        loc = url_tag.find('loc').text if url_tag.find('loc') else None
        lastmod = url_tag.find('lastmod').text if url_tag.find('lastmod') else None
        urls_data.append((loc, lastmod))
    return urls_data

def save_to_excel(new_urls):
    if not os.path.exists(EXCEL_FILE_NAME):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["URL", "Last Modified", "Time when URL found"])
    else:
        workbook = load_workbook(EXCEL_FILE_NAME)
        worksheet = workbook.active

    for url, lastmod in new_urls:
        worksheet.append([url, lastmod, current_timestamp()])
    
    workbook.save(EXCEL_FILE_NAME)

def get_saved_urls():
    if not os.path.exists(EXCEL_FILE_NAME):
        return []
    
    workbook = load_workbook(EXCEL_FILE_NAME, read_only=True)
    worksheet = workbook.active

    return [row[0].value for row in list(worksheet.iter_rows())[1:]]

def main():
    while True:
        print(f"[{current_timestamp()}] - XML content ophalen...")
        xml_content = fetch_xml_content(URL)
        current_urls_data = extract_urls_and_lastmod(xml_content)
        saved_urls = get_saved_urls()

        new_urls = [(url, lastmod) for url, lastmod in current_urls_data if url not in saved_urls]

        if new_urls:
            print(f"[{current_timestamp()}] - Nieuwe URLs gevonden, ik sla het op in Excel...")
            save_to_excel(new_urls)
        else:
            print(f"[{current_timestamp()}] - Geen nieuwe URLs gevonden.")
        
        print(f"[{current_timestamp()}] - Ik check over 4 uur weer...")
        time.sleep(SLEEP_INTERVAL)

if __name__ == "__main__":
    main()
